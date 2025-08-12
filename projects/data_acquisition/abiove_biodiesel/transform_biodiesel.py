import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))                

from typing import List
import pandas as pd 
from projects.core.data_acquisition import TransformData
from projects.data_acquisition.abiove_biodiesel.data_contract import biodiesel_schema
from projects.data_acquisition.abiove_biodiesel.constants_biodiesel import ITF, INTERVALO, FAMCOML, SHEETS_INFOS, MONTH_DICT, SOURCE_NAME
import openpyxl
from openpyxl import load_workbook
from io import BytesIO

class TransformBiodiesel(TransformData):
    def __init__(self) -> None:
        super().__init__()
        
    def transform_data(self, blob_path: List[str], exec_date) -> List[pd.DataFrame]:
        """
        Função que transforma os arquivos brutos de biodiesel da Abiove
        
        :param blob_path: lista com os paths dos arquivos brutos
        
        :return: lista com os dataframes transformados
        
        """
        
        file_content_list = []
        for blob_name in blob_path:
            raw_data = self.download_raw_data(raw_data_path=blob_name)
            
            file_content_list.append(raw_data)
            
        consolidated_file = self._consolidate_raw_workbook(file_content_list=file_content_list)
        
        df_concated = self._transform_each_sheet(raw_data=consolidated_file, execution_date=exec_date)
        
        df_cleaned = self.clean_data_up(dataframe=df_concated)
        
        df_transformed = self._transform_dataframe_to_default_layout(dataframe=df_cleaned, itf=ITF, famcompl=FAMCOML, intervalo=INTERVALO, original_columns=['Value'], attributes_columns=['a24'])

        output_data = self.create_output_dict(execution_date=exec_date, dataframe=df_transformed, identifier=ITF, schema=biodiesel_schema)

        output_dict = self.upload_output(output_data=output_data)
        
        return output_dict
            
    def _transform_each_sheet(self, raw_data: pd.DataFrame, execution_date) -> pd.DataFrame:
        """
        Função que transforma cada planilha do arquivo consolidado de biodiesel da Abiove
        
        :param raw_data: arquivo consolidado de biodiesel da Abiove
        
        :return: dataframe com todas as planilhas transformadas
        """
        
        df_concat = pd.DataFrame()
        
        for sheet in SHEETS_INFOS:
            sheet_name = sheet['sheet_name']
            serie_name = sheet['serie_name']
                        
            if sheet_name == 'producao_m3_total':
                df_temp = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=6, skip_footer=3)
                
                df_transformed = self._transform_raw_dataframe(df_temp=df_temp, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}', sheet_name=sheet_name)
                
                df_concat = pd.concat([df_concat, df_transformed], axis=0)

            elif sheet_name == 'producao_m3_regiao':
                df_temp = pd.read_excel(raw_data, sheet_name=sheet_name, skiprows=32, skipfooter=5)
                                
                df_transformed = self._transform_raw_dataframe(df_temp=df_temp, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}', sheet_name=sheet_name)

                df_concat = pd.concat([df_concat, df_transformed], axis=0)
                
            elif sheet_name == 'entrega_venda_m3':
                df_temp = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=6, skip_footer=5)
                
                df_transformed = self._transform_raw_dataframe(df_temp=df_temp, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}', sheet_name=sheet_name)
                               
                df_concat = pd.concat([df_concat, df_transformed], axis=0)
            

            elif sheet_name == 'materia-prima_anual':
                df_temp1 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=6, skip_footer=16)
                
                df_transformed1 = self._transform_raw_dataframe(df_temp=df_temp1, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table1', sheet_name=sheet_name)
                
                df_temp2 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=19, skip_footer=3)

                df_transformed2 = self._transform_raw_dataframe(df_temp=df_temp2, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table2', sheet_name=sheet_name)
                
                df_concat = pd.concat([df_concat, df_transformed1, df_transformed2], axis=0)
            
            
            elif sheet_name == 'vendas_importacao_dieselB':
                df_temp1 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=6, skip_footer=76)
                
                df_transformed1 = self._transform_raw_dataframe(df_temp=df_temp1, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table1', sheet_name=sheet_name)
                
                df_temp2 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=25, skip_footer=57)
                
                df_transformed2 = self._transform_raw_dataframe(df_temp=df_temp2, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table2', sheet_name=sheet_name)
                
                df_temp3 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=43, skip_footer=39)
                
                df_transformed3 = self._transform_raw_dataframe(df_temp=df_temp3, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table3', sheet_name=sheet_name)
                
                df_temp4 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=61, skip_footer=21)
                
                df_transformed4 = self._transform_raw_dataframe(df_temp=df_temp4, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table4', sheet_name=sheet_name)
                
                df_temp5 = self._read_biodiesel_excel_files(raw_data=raw_data, sheet_name=sheet_name, skip_rows=80, skip_footer=2)
                
                df_transformed5 = self._transform_raw_dataframe(df_temp=df_temp5, exec_date=execution_date, serie_name=serie_name, file_name=f'{sheet_name}_table5', sheet_name=sheet_name)
                
                df_concat = pd.concat([df_concat, df_transformed1, df_transformed2, df_transformed3, df_transformed4, df_transformed5], axis=0)
                
            
            return df_concat

    def _reshape_dataframe(self, dataframe):
        """
        Função que faz o reshape do dataframe
        
        :param dataframe: dataframe a ser transformado
        
        :return: dataframe transformado
        """
        column_last_year = dataframe.columns[-1]
        first_column = dataframe.columns[0]
        
        dataframe[first_column] = dataframe[first_column].replace(MONTH_DICT)
        
        df_filtred = dataframe[[first_column, column_last_year]]

        df_melted = df_filtred.melt(id_vars=[first_column], var_name='Year', value_name='Value')

        df_melted['DATASER'] = df_melted.apply(lambda row: f"{row['Year']}" + '-' + f"{row[first_column]}" + '-' + '01', axis=1)
        
        df_melted = df_melted[['DATASER', 'Value']]

        return df_melted


    def _read_biodiesel_excel_files(self, raw_data, sheet_name: str, skip_rows: int, skip_footer: int):
        """
        Função que lê os arquivos de biodiesel da Abiove
        
        :param raw_data: arquivo de biodiesel da Abiove
        
        :param sheet_name: nome da planilha
        
        :param skip_rows: número de linhas a serem puladas no início da planilha
        
        :param skip_footer: número de linhas a serem puladas no final da planilha
        
        :return: dataframe com a planilha lida
        """
        df_temp = pd.read_excel(raw_data, sheet_name=sheet_name, skiprows=skip_rows, skipfooter=skip_footer)
        df_temp = df_temp.drop(df_temp.columns[0], axis=1)
        
        return df_temp
    
    def _transform_sheet_with_differents_series_name(self, dataframe: pd.DataFrame, serie_name: str, dataser):
        """
        Função que transforma as planilhas que possuem séries com nomes diferentes
        
        :param dataframe: dataframe a ser transformado
        
        :param serie_name: nome da série
        
        :param dataser: data de execução da extração
        
        :return: dataframe transformado
        """
    
        first_column = dataframe.columns[0]
        last_column = dataframe.columns[-1]
        
        dataframe['SERIE'] = dataframe.apply(lambda row: serie_name + ' - ' + f"{row[first_column]}", axis=1)
        dataframe['DATASER'] = dataser.strf('%Y-%m-%d')
        
        dataframe.rename(columns={last_column: 'Value'}, inplace=True)
        
        df_transformed = dataframe[['DATASER', 'Value', 'SERIE']]
        
        return df_transformed
    
    
    def _transform_raw_dataframe(self, df_temp, exec_date, serie_name, file_name, sheet_name):
        """
        Função que transforma o dataframe bruto
        
        :param df_temp: dataframe bruto
        
        :param exec_date: data de execução da extração
        
        :param serie_name: nome da série
        
        :param file_name: nome do arquivo
        
        :param sheet_name: nome da planilha
        
        :return: dataframe transformado
        """
        
        different_sheet_list = ['producao_m3_regiao', 'materia-prima_anual']
        
        cleaned_df = self.clean_data_up(dataframe=df_temp, source_name=SOURCE_NAME, dataser=exec_date, file_name=file_name)
        
        if sheet_name in different_sheet_list:
            df_transformed = self._transform_sheet_with_differents_series_name(dataframe=cleaned_df, serie_name=serie_name, dataser=exec_date)

        else:
            df_transformed = self._reshape_dataframe(dataframe=cleaned_df)
            
            df_transformed['SERIE'] = serie_name
            
        return df_transformed
    
    def _consolidate_raw_workbook(self, file_content_list: List[BytesIO]):
        """
        Função que consolida os arquivos brutos de biodiesel da Abiove
        
        :param file_content_list: lista com os arquivos brutos
        
        :return: BytesIO com o arquivo consolidado
        """
        consolidated_workbook = openpyxl.Workbook()

        consolidated_workbook.remove(consolidated_workbook.active)

        for index, file_content in enumerate(file_content_list):
            if isinstance(file_content, BytesIO):
                workbook = load_workbook(file_content)
            else:
                raise ValueError(f"Invalid file content type: {type(file_content)}")

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                new_sheet = consolidated_workbook.create_sheet(title=f'{sheet_name}')
                
                for row in sheet.iter_rows(values_only=True):
                    new_sheet.append(row)
                    
        excel_bytes = BytesIO()
        consolidated_workbook.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes
